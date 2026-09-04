#!/usr/bin/env python3
"""Append M. Hafiz's 30 Juli 2026 Batin XXIV survey into points.geojson.

Source: the Timemark "Lembar Foto" export LembarFoto_2026-07-24_to_2026-07-30.xlsx
(ESDM/PJUTS/OLAK BESAR/RAW). Its "Foto" sheet holds 37 photos embedded in
column A, one per row, every row dated 2026-07-30 in Kec. Batin XXIV, Kab.
Batang Hari. The surveyor typed the point number into "Nama" ("01", "02 Karmeo",
"05 Jangga Baru"); Olak Besar and Durian Luncuk rows carry only the number, so
the desa comes from the Timemark geocode in "Lokasi" and, for the three Olak
Besar rows geocoded as "Unnamed Road, Kec. Batin XXIV", from the visit order.

The 37 photos describe 20 points, which is exactly the Batin XXIV share of
rekapan A.1 (M. HAFIZ, jalur Ketua DPRD): Durian Luncuk 5, Jangga Baru 5, Olak
Besar 4, Karmeo 4, Desa Simp. Aur Gading 2. 13 of the 37 photos are not point
photos at all: arriving at the kantor desa and signing the berita acara (Karmeo
14:22-14:23, Jangga Baru 16:18-16:30) and the crew's vehicles/selfies at Olak
Besar (11:03-11:04). They inherit whichever "Nama" was active, so the earliest
photo of a label is *not* always the point. Each point is therefore pinned to
an explicit source row below - the earliest photo that shows the spot itself -
and takes the coordinate stamped on that photo.

Olak Besar 003 (11:17:43) and 004 (11:20:13) carry the identical stamp
(-1.89677, 102.98033, 29.0 m) although the pictures show different places; the
phone GPS did not refresh in between. Both rows are flagged Duplikat, as with
Sungai Ulas 001/002, so the crew knows the two pins are provisional.

Photos exported at 720x1087 carry a white caption strip below the picture with
the Timemark album name "DESA OLAK BESAR" - also on Karmeo, Durian Luncuk, Aur
Gading and Jangga Baru photos. The strip is cropped away so the atlas photo is
the 720x960 picture, matching the other Timemark photos in images/.
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from copy import deepcopy
from pathlib import Path

import openpyxl
from PIL import Image

FOTO_PREFIX = "D:/011. ESDM/Penerangan Jalan Umum Tenaga Surya 2026/2026/03. Foto"
TANGGAL = "30/07/2026"
NAMA = "M. Hafiz"
JALUR = "Ketua DPRD"
KAB = "Batang Hari"
KEC = "Batin XXIV"

# Nomor desa -> (desa label for Alamat, sheet rows that belong to the desa).
# Rows are the Foto sheet rows (header is row 1); listed for the whole desa so
# the script can verify every source row is accounted for exactly once.
DESA: dict[str, tuple[str, list[int]]] = {
    "OLAK BESAR": ("Desa Olak Besar", [31, 32, 33, 34, 35, 36, 37, 38]),
    "DURIAN LUNCUK": ("Desa Durian Luncuk", [24, 25, 26, 27, 28, 29, 30]),
    "SIMPANG AUR GADING": ("Desa Simpang Aur Gading", [22, 23]),
    "KARMEO": ("Desa Karmeo", [14, 15, 16, 17, 18, 19, 20, 21]),
    "JANGGA BARU": ("Desa Jangga Baru", [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]),
}

# (Nomor desa, seq) -> (source row, expected Waktu, expected label number,
# Keterangan, duplikat). The row is the photo that shows the spot; the Waktu
# and label are asserted against the sheet so a re-export cannot silently
# shift the mapping.
POINTS: list[tuple[str, int, int, str, str, str, bool]] = [
    ("OLAK BESAR", 1, 38, "10:59:50", "01", "RT 02 Olak Besar", False),
    # Rows 34-37 (11:03-11:04, label 02) are the crew's motorbike, car and a
    # selfie 10 m from 001; the point itself is the 11:15 photo 358 m south.
    ("OLAK BESAR", 2, 33, "11:15:23", "02", "Desa Olak Besar", False),
    ("OLAK BESAR", 3, 32, "11:17:43", "03", "Desa Olak Besar", True),
    ("OLAK BESAR", 4, 31, "11:20:13", "04", "Desa Olak Besar", True),
    ("DURIAN LUNCUK", 1, 30, "12:08:53", "01", "Desa Durian Luncuk", False),
    ("DURIAN LUNCUK", 2, 29, "12:11:04", "02", "Desa Durian Luncuk", False),
    ("DURIAN LUNCUK", 3, 27, "12:28:10", "03", "Desa Durian Luncuk", False),
    ("DURIAN LUNCUK", 4, 25, "12:34:26", "04", "Desa Durian Luncuk", False),
    ("DURIAN LUNCUK", 5, 24, "12:42:38", "05", "Desa Durian Luncuk", False),
    ("SIMPANG AUR GADING", 1, 23, "13:16:30", "01", "Desa Simpang Aur Gading", False),
    ("SIMPANG AUR GADING", 2, 22, "13:17:59", "02", "Desa Simpang Aur Gading", False),
    # Rows 19-21 (14:22-14:23, label 01) are the berita acara being signed on
    # a porch in RT 08, 8 m from point 004; the point is the 14:35 field photo.
    ("KARMEO", 1, 18, "14:35:30", "01", "Desa Karmeo", False),
    ("KARMEO", 2, 16, "14:38:17", "02", "Desa Karmeo", False),
    ("KARMEO", 3, 15, "14:43:43", "03", "Desa Karmeo", False),
    ("KARMEO", 4, 14, "14:48:19", "04", "RT 08 Karmeo", False),
    # Rows 8-13 (16:18-16:30, label 01) are the arrival at the kantor desa and
    # the meeting inside; the point is the 16:43 photo on the road, 38 m away.
    ("JANGGA BARU", 1, 7, "16:43:01", "01", "Desa Jangga Baru", False),
    ("JANGGA BARU", 2, 6, "16:47:48", "02", "Desa Jangga Baru", False),
    ("JANGGA BARU", 3, 4, "16:56:25", "03", "Desa Jangga Baru", False),
    ("JANGGA BARU", 4, 3, "17:02:56", "04", "Desa Jangga Baru", False),
    ("JANGGA BARU", 5, 2, "17:08:39", "05", "Desa Jangga Baru", False),
]

EXPECTED_PHOTOS = 37
EXPECTED_POINTS = 20
PHOTO_SIZE = (720, 960)  # Timemark portrait frame without the caption strip

COORD_RE = re.compile(r"^(-?\d+\.\d+)°S,(-?\d+\.\d+)°E")


def parse_coord(text: str) -> tuple[float, float]:
    """'-1.89282°S,102.98038°E ' -> (lon, lat)."""
    m = COORD_RE.match(str(text).strip())
    if not m:
        raise SystemExit(f"Unparseable Koordinat: {text!r}")
    lat, lon = float(m.group(1)), float(m.group(2))
    return lon, lat


def sanitize_media_path(value: str) -> str:
    return re.sub(r"[\\/:]", "_", str(value or "")).strip()


def crop_caption(data: bytes) -> bytes:
    """Drop the white album-name strip Timemark adds below 720x1087 exports."""
    im = Image.open(io.BytesIO(data))
    if im.size == PHOTO_SIZE:
        return data
    if im.size != (720, 1087):
        raise SystemExit(f"Unexpected photo size {im.size}")
    strip = im.convert("L").crop((0, PHOTO_SIZE[1], 720, 1087))
    # The strip is white with a few dark caption glyphs. Refuse to crop if the
    # dark share is large enough to be picture content instead.
    dark = sum(strip.histogram()[:120])
    if dark > strip.size[0] * strip.size[1] * 0.05:
        raise SystemExit("Bottom strip does not look like a caption")
    out = io.BytesIO()
    im.convert("RGB").crop((0, 0, 720, PHOTO_SIZE[1])).save(out, "JPEG", quality=92)
    return out.getvalue()


def read_timemark(xlsx_path: Path) -> dict[int, dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Foto"]
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}
    for needed in ("Tanggal", "Waktu", "Lokasi", "Nama", "Koordinat"):
        if needed not in col:
            raise SystemExit(f"Column {needed!r} missing from Foto sheet")

    photos_by_row: dict[int, bytes] = {}
    for img in ws._images:  # anchors tie each photo to its row
        row = img.anchor._from.row + 1
        if row in photos_by_row:
            raise SystemExit(f"Two photos anchored on row {row}")
        photos_by_row[row] = img._data()
    if len(photos_by_row) != EXPECTED_PHOTOS:
        raise SystemExit(f"Expected {EXPECTED_PHOTOS} embedded photos, found {len(photos_by_row)}")

    rows: dict[int, dict] = {}
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, col["Nama"]).value
        if label is None:
            continue
        tanggal = str(ws.cell(row, col["Tanggal"]).value)
        if tanggal != "2026-07-30":
            raise SystemExit(f"Row {row}: unexpected Tanggal {tanggal!r}")
        if row not in photos_by_row:
            raise SystemExit(f"Row {row} ({label}) has no embedded photo")
        lon, lat = parse_coord(ws.cell(row, col["Koordinat"]).value)
        num = re.search(r"\d+", str(label))
        rows[row] = {
            "label": str(label).strip(),
            "num": num.group(0) if num else "",
            "waktu": str(ws.cell(row, col["Waktu"]).value),
            "lokasi": str(ws.cell(row, col["Lokasi"]).value),
            "lon": lon,
            "lat": lat,
            "photo": photos_by_row[row],
        }
    if len(rows) != EXPECTED_PHOTOS:
        raise SystemExit(f"Expected {EXPECTED_PHOTOS} rows with a label, found {len(rows)}")
    return rows


def check_desa_map(rows: dict[int, dict]) -> None:
    seen: list[int] = []
    for desa, (_, desa_rows) in DESA.items():
        seen.extend(desa_rows)
        for r in desa_rows:
            lokasi = rows[r]["lokasi"].lower()
            label = rows[r]["label"].lower()
            key = desa.lower().replace("simpang ", "")
            if key not in lokasi and key not in label and "unnamed road, kec. batin xxiv" not in lokasi:
                raise SystemExit(f"Row {r} ({rows[r]['label']!r}, {rows[r]['lokasi']!r}) does not read as {desa}")
    if sorted(seen) != sorted(rows):
        raise SystemExit("DESA rows do not cover the sheet exactly once")


def build_features(rows: dict[int, dict], first_fid: int) -> list[dict]:
    features: list[dict] = []
    used: set[int] = set()
    for desa_code, seq, row, waktu, num, keterangan, duplikat in POINTS:
        desa_label, desa_rows = DESA[desa_code]
        src = rows[row]
        if row not in desa_rows:
            raise SystemExit(f"Row {row} is not a {desa_code} row")
        if src["waktu"] != waktu or src["num"] != num:
            raise SystemExit(
                f"Row {row}: expected {waktu} label {num}, sheet has {src['waktu']} label {src['num']}"
            )
        if row in used:
            raise SystemExit(f"Row {row} used twice")
        used.add(row)
        nomor = f"{KAB.upper()}-{KEC.upper()}-{desa_code}-{seq:03d}"
        props = {
            "fid": str(first_fid + len(features)),
            "Nomor": nomor,
            "Nama Anggota": NAMA,
            "Jalur": JALUR,
            "Alamat": f"{desa_label}, Kecamatan {KEC}, Kabupaten {KAB}",
            "Longitude": src["lon"],
            "Latitude": src["lat"],
            "Tanggal Dokumentasi": TANGGAL,
            "Keterangan": keterangan,
            "Foto Survey Awal": f"{FOTO_PREFIX}/{nomor}.jpg",
        }
        if duplikat:
            props["Duplikat"] = True
        features.append(
            {
                "type": "Feature",
                "properties": props,
                "geometry": {"type": "Point", "coordinates": [src["lon"], src["lat"]]},
                "_photo": crop_caption(src["photo"]),
                "_row": row,
                "_shots": sum(1 for r in desa_rows if rows[r]["num"] == num),
            }
        )
    if len(features) != EXPECTED_POINTS:
        raise SystemExit(f"Expected {EXPECTED_POINTS} points, built {len(features)}")

    # Rekapan A.1 share for Batin XXIV.
    per_desa = {d: sum(1 for f in features if f["properties"]["Nomor"].split("-")[2] == d) for d in DESA}
    if per_desa != {"OLAK BESAR": 4, "DURIAN LUNCUK": 5, "SIMPANG AUR GADING": 2, "KARMEO": 4, "JANGGA BARU": 5}:
        raise SystemExit(f"Per-desa counts do not match the rekapan: {per_desa}")

    # The Duplikat pair must really share a stamp; nothing else may.
    coords: dict[tuple[float, float], list[str]] = {}
    for f in features:
        coords.setdefault(tuple(f["geometry"]["coordinates"]), []).append(f["properties"]["Nomor"])
    for c, nomors in coords.items():
        flagged = all(
            next(f for f in features if f["properties"]["Nomor"] == n)["properties"].get("Duplikat")
            for n in nomors
        )
        if len(nomors) > 1 and not flagged:
            raise SystemExit(f"Shared coordinate without Duplikat flag: {nomors}")
        if len(nomors) == 1 and flagged:
            raise SystemExit(f"Duplikat flag on a unique coordinate: {nomors}")
    return features


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("/Users/dany/ESDM/PJUTS/OLAK BESAR/RAW/LembarFoto_2026-07-24_to_2026-07-30.xlsx"),
    )
    parser.add_argument("--repo", type=Path, default=Path(__file__).resolve().parent.parent)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    points_path = args.repo / "data" / "points.geojson"
    images_dir = args.repo / "images"
    for path in (args.xlsx, points_path, images_dir):
        if not path.exists():
            raise SystemExit(f"Missing input: {path}")

    existing = json.loads(points_path.read_text(encoding="utf-8"))
    old_features = existing["features"]
    old_snapshot = deepcopy(old_features)
    old_nomors = {f["properties"]["Nomor"] for f in old_features}
    max_fid = max(int(f["properties"]["fid"]) for f in old_features)

    rows = read_timemark(args.xlsx)
    check_desa_map(rows)
    features = build_features(rows, max_fid + 1)

    for f in features:
        nomor = f["properties"]["Nomor"]
        if nomor in old_nomors:
            raise SystemExit(f"Nomor already in atlas: {nomor}")
        dst = images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])
        if dst.exists():
            raise SystemExit(f"Photo already present: {dst.name}")

    print(f"existing: {len(old_features)}  append: {len(features)}  total: {len(old_features) + len(features)}")
    for f in features:
        p = f["properties"]
        flag = "  DUPLIKAT" if p.get("Duplikat") else ""
        print(
            f"  {p['Nomor']:<44} {p['Latitude']:>9} {p['Longitude']:>10}  "
            f"row {f['_row']:>2} of {f['_shots']}  {rows[f['_row']]['waktu']}  {p['Keterangan']}{flag}"
        )

    if args.dry_run:
        print("\ndry-run: nothing written")
        return 0

    for f in features:
        dst = images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])
        dst.write_bytes(f.pop("_photo"))
        f.pop("_row")
        f.pop("_shots")

    merged = dict(existing)
    merged["features"] = old_features + features
    points_path.write_text(
        json.dumps(merged, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    written = json.loads(points_path.read_text(encoding="utf-8"))
    if written["features"][: len(old_snapshot)] != old_snapshot:
        raise SystemExit("Old features changed after write")
    if len(written["features"]) != len(old_features) + len(features):
        raise SystemExit("Feature count mismatch after write")
    nomors = [f["properties"]["Nomor"] for f in written["features"]]
    if len(set(nomors)) != len(nomors):
        raise SystemExit("Duplicate Nomor after write")
    for f in features:
        dst = images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])
        if not dst.is_file():
            raise SystemExit(f"Photo not written for {f['properties']['Nomor']}")
        if Image.open(dst).size != PHOTO_SIZE:
            raise SystemExit(f"Photo {dst.name} is not {PHOTO_SIZE}")

    print(f"\nOK  photos written: {len(features)}  total features: {len(written['features'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
