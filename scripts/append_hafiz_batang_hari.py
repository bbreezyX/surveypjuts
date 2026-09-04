#!/usr/bin/env python3
"""Append M. Hafiz's 30 Juli 2026 Batang Hari survey into points.geojson.

Source: a Timemark "Lembar Foto" export (LembarFoto_2026-08-13_to_2026-08-19.xlsx)
whose "Foto" sheet holds 48 photos embedded in column A, one per row, with the
surveyor's title in "Judul" and the GPS stamp in "Koordinat". Every row is
dated 2026-07-30 and every photo carries the same stamp burned in.

The 48 photos describe 24 points. The surveyor shot each point twice (three
times for one): a close-up of the ground and, a few seconds earlier, a wide
shot with the people standing at the spot. The wide shot is the earliest photo
of each group and is the one worth showing in the popup, so that is the one
copied; the point takes the coordinate stamped on that same photo so the pin
never disagrees with the picture.

Rekapan A.1 (M. HAFIZ, jalur Ketua DPRD) lists these Muara Bulian / Pemayung
rows: Pesantren Al-Muhajirin 4, Malapari 4, Perumnas 2, Pondok Indah RT 36 2,
RT 15 Rengas Condong 2, RT 14 Rengas Condong 2, RT 11 Teratai (Mayang
Mangurai) 3, RT 21 Rengas Condong 3, Ture 2, plus RT 06 Teratai 2 and RT 10
Sei. Bulu 4 that this survey did not visit. Desa/kelurahan in the Nomor follow
the GPS stamp and the surveyor's title, as with Sungai Ulas and Tempino:
Al-Muhajirin sits in Desa Simpang Terusan (its own address says so) even though
the rekapan writes Tenam, and the "RT 15" points are stamped Kel. Teratai
although the rekapan puts RT 15 under Rengas Condong.

Ture already has TURE-001..010 (Sapuan Ansori), so Hafiz's two continue as
011 and 012 rather than starting a second sequence for the same desa.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from copy import deepcopy
from pathlib import Path

import openpyxl

FOTO_PREFIX = "D:/011. ESDM/Penerangan Jalan Umum Tenaga Surya 2026/2026/03. Foto"
TANGGAL = "30/07/2026"
NAMA = "M. Hafiz"
JALUR = "Ketua DPRD"
KAB = "Batang Hari"

# Surveyor title (lower-cased) -> (kecamatan, desa/kel label, Nomor desa, sequence, Keterangan).
# Sequence numbers follow the order the surveyor visited the spots.
TITLES: dict[str, tuple[str, str, str, int, str]] = {
    "pesantren al muhajirin 01": ("Muara Bulian", "Desa Simpang Terusan", "SIMPANG TERUSAN", 1, "Pesantren Al-Muhajirin 01"),
    "pesantren al muhajirin 02": ("Muara Bulian", "Desa Simpang Terusan", "SIMPANG TERUSAN", 2, "Pesantren Al-Muhajirin 02"),
    "pesantren al muhajirin 03": ("Muara Bulian", "Desa Simpang Terusan", "SIMPANG TERUSAN", 3, "Pesantren Al-Muhajirin 03"),
    "pesantren al muhajirin 04": ("Muara Bulian", "Desa Simpang Terusan", "SIMPANG TERUSAN", 4, "Pesantren Al-Muhajirin 04"),
    "desa malapari 01": ("Muara Bulian", "Desa Malapari", "MALAPARI", 1, "Desa Malapari 01"),
    "desa malapari 02": ("Muara Bulian", "Desa Malapari", "MALAPARI", 2, "Desa Malapari 02"),
    "desa malapari 03": ("Muara Bulian", "Desa Malapari", "MALAPARI", 3, "Desa Malapari 03"),
    "desa malapari 04": ("Muara Bulian", "Desa Malapari", "MALAPARI", 4, "Desa Malapari 04"),
    "perumnas 01": ("Muara Bulian", "Kel. Muara Bulian", "MUARA BULIAN", 1, "Perumnas 01"),
    "perumnas 02": ("Muara Bulian", "Kel. Muara Bulian", "MUARA BULIAN", 2, "Perumnas 02"),
    "perumahan pondok berlian indah 01": ("Muara Bulian", "Kel. Muara Bulian", "MUARA BULIAN", 3, "Perumahan Pondok Berlian Indah 01"),
    "perumahan pondok berlian indah 02": ("Muara Bulian", "Kel. Muara Bulian", "MUARA BULIAN", 4, "Perumahan Pondok Berlian Indah 02"),
    "rt. 21 rengas condong": ("Muara Bulian", "Kel. Rengas Condong", "RENGAS CONDONG", 1, "RT 21 Rengas Condong 01"),
    "rt. 21 rengas condong 02": ("Muara Bulian", "Kel. Rengas Condong", "RENGAS CONDONG", 2, "RT 21 Rengas Condong 02"),
    "rt. 14 rengas condong 01": ("Muara Bulian", "Kel. Rengas Condong", "RENGAS CONDONG", 3, "RT 14 Rengas Condong 01"),
    "rt. 14 rengas condong 02": ("Muara Bulian", "Kel. Rengas Condong", "RENGAS CONDONG", 4, "RT 14 Rengas Condong 02"),
    "rt. 15 teratai 01": ("Muara Bulian", "Kel. Teratai", "TERATAI", 1, "RT 15 Teratai 01"),
    "rt. 15 teratai 02": ("Muara Bulian", "Kel. Teratai", "TERATAI", 2, "RT 15 Teratai 02"),
    "rt. 15 teratai 03": ("Muara Bulian", "Kel. Teratai", "TERATAI", 3, "RT 15 Teratai 03"),
    "mayang mangurai 01": ("Muara Bulian", "Kel. Teratai", "TERATAI", 4, "Mayang Mangurai 01"),
    "mayang mangurai 02": ("Muara Bulian", "Kel. Teratai", "TERATAI", 5, "Mayang Mangurai 02"),
    "mayang mangurai 03": ("Muara Bulian", "Kel. Teratai", "TERATAI", 6, "Mayang Mangurai 03"),
    "desa ture 01": ("Pemayung", "Desa Ture", "TURE", 11, "Desa Ture 01"),
    "desa ture 02": ("Pemayung", "Desa Ture", "TURE", 12, "Desa Ture 02"),
}

EXPECTED_PHOTOS = 48
EXPECTED_POINTS = 24

COORD_RE = re.compile(r"^(-?\d+\.\d+)°S,(-?\d+\.\d+)°E")


def parse_coord(text: str) -> tuple[float, float]:
    """'-1.56205°S,103.41671°E ±52m' -> (lon, lat)."""
    m = COORD_RE.match(str(text).strip())
    if not m:
        raise SystemExit(f"Unparseable Koordinat: {text!r}")
    lat, lon = float(m.group(1)), float(m.group(2))
    return lon, lat


def sanitize_media_path(value: str) -> str:
    return re.sub(r"[\\/:]", "_", str(value or "")).strip()


def read_timemark(xlsx_path: Path) -> "OrderedDict[str, list[dict]]":
    """Group the Foto sheet rows by title, each row carrying its embedded photo."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Foto"]
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}
    for needed in ("Tanggal", "Waktu", "Judul", "Koordinat"):
        if needed not in col:
            raise SystemExit(f"Column {needed!r} missing from Foto sheet")

    photos_by_row: dict[int, bytes] = {}
    for img in ws._images:  # openpyxl keeps anchors; the row is what ties a photo to its text
        row = img.anchor._from.row + 1
        if row in photos_by_row:
            raise SystemExit(f"Two photos anchored on row {row}")
        photos_by_row[row] = img._data()
    if len(photos_by_row) != EXPECTED_PHOTOS:
        raise SystemExit(f"Expected {EXPECTED_PHOTOS} embedded photos, found {len(photos_by_row)}")

    groups: "OrderedDict[str, list[dict]]" = OrderedDict()
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, col["Judul"]).value
        if not title:
            continue
        tanggal = str(ws.cell(row, col["Tanggal"]).value)
        if tanggal != "2026-07-30":
            raise SystemExit(f"Row {row}: unexpected Tanggal {tanggal!r}")
        if row not in photos_by_row:
            raise SystemExit(f"Row {row} ({title}) has no embedded photo")
        lon, lat = parse_coord(ws.cell(row, col["Koordinat"]).value)
        groups.setdefault(str(title).strip().lower(), []).append(
            {
                "row": row,
                "waktu": str(ws.cell(row, col["Waktu"]).value),
                "lon": lon,
                "lat": lat,
                "photo": photos_by_row[row],
            }
        )
    return groups


def build_features(groups: "OrderedDict[str, list[dict]]", first_fid: int) -> list[dict]:
    unknown = sorted(set(groups) - set(TITLES))
    missing = sorted(set(TITLES) - set(groups))
    if unknown or missing:
        raise SystemExit(f"Title mismatch.\n  not in table: {unknown}\n  not in sheet: {missing}")

    features: list[dict] = []
    for title, (kec, desa_label, desa_code, seq, keterangan) in TITLES.items():
        shots = sorted(groups[title], key=lambda s: s["waktu"])
        chosen = shots[0]  # earliest = wide shot with people
        nomor = f"{KAB.upper()}-{kec.upper()}-{desa_code}-{seq:03d}"
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "fid": str(first_fid + len(features)),
                    "Nomor": nomor,
                    "Nama Anggota": NAMA,
                    "Jalur": JALUR,
                    "Alamat": f"{desa_label}, Kecamatan {kec}, Kabupaten {KAB}",
                    "Longitude": chosen["lon"],
                    "Latitude": chosen["lat"],
                    "Tanggal Dokumentasi": TANGGAL,
                    "Keterangan": keterangan,
                    "Foto Survey Awal": f"{FOTO_PREFIX}/{nomor}.jpg",
                },
                "geometry": {"type": "Point", "coordinates": [chosen["lon"], chosen["lat"]]},
                "_photo": chosen["photo"],
                "_row": chosen["row"],
                "_shots": len(shots),
            }
        )
    if len(features) != EXPECTED_POINTS:
        raise SystemExit(f"Expected {EXPECTED_POINTS} points, built {len(features)}")
    return features


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("/Users/dany/ESDM/PJUTS/LembarFoto_2026-08-13_to_2026-08-19.xlsx"),
    )
    parser.add_argument("--repo", type=Path, default=Path(__file__).resolve().parent.parent)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    points_path = args.repo / "data" / "points.geojson"
    images_dir = args.repo / "images"
    for path in (args.xlsx, points_path, images_dir):
        if not path.exists():
            raise SystemExit(f"Missing input: {path}")

    existing = json.loads(points_path.read_text(encoding="utf-8"))
    old_features = existing["features"]
    old_snapshot = deepcopy(old_features)
    old_nomors = {f["properties"]["Nomor"] for f in old_features}
    max_fid = max(int(f["properties"]["fid"]) for f in old_features)

    groups = read_timemark(args.xlsx)
    features = build_features(groups, max_fid + 1)

    for f in features:
        nomor = f["properties"]["Nomor"]
        if nomor in old_nomors:
            raise SystemExit(f"Nomor already in atlas: {nomor}")
        dst = images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])
        if dst.exists():
            raise SystemExit(f"Photo already present: {dst.name}")

    print(f"existing: {len(old_features)}  append: {len(features)}  total: {len(old_features) + len(features)}")
    for f in features:
        p = f["properties"]
        print(
            f"  {p['Nomor']:<48} {p['Latitude']:>9} {p['Longitude']:>10}  "
            f"row {f['_row']:>2} of {f['_shots']}  {p['Keterangan']}"
        )

    if args.dry_run:
        print("\ndry-run: nothing written")
        return 0

    for f in features:
        dst = images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])
        dst.write_bytes(f.pop("_photo"))
        f.pop("_row")
        f.pop("_shots")

    merged = dict(existing)
    merged["features"] = old_features + features
    points_path.write_text(
        json.dumps(merged, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    written = json.loads(points_path.read_text(encoding="utf-8"))
    if written["features"][: len(old_snapshot)] != old_snapshot:
        raise SystemExit("Old features changed after write")
    if len(written["features"]) != len(old_features) + len(features):
        raise SystemExit("Feature count mismatch after write")
    nomors = [f["properties"]["Nomor"] for f in written["features"]]
    if len(set(nomors)) != len(nomors):
        raise SystemExit("Duplicate Nomor after write")
    for f in features:
        if not (images_dir / sanitize_media_path(f["properties"]["Foto Survey Awal"])).is_file():
            raise SystemExit(f"Photo not written for {f['properties']['Nomor']}")

    print(f"\nOK  photos written: {len(features)}  total features: {len(written['features'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
